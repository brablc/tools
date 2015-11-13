#!/usr/bin/env python
"""
FUNCTION: Converts a CSV (tab delimited) file to an Excel xlsx file.

Copyright (c) 2012, Konrad Foerstner <konrad@foerstner.org>
Copyright (c) 2013, Ondrej Brablc <ondrej@brablc.com>

Permission to use, copy, modify, and/or distribute this software for
any purpose with or without fee is hereby granted, provided that the
above copyright notice and this permission notice appear in all
copies.

THE SOFTWARE IS PROVIDED 'AS IS' AND THE AUTHOR DISCLAIMS ALL
WARRANTIES WITH REGARD TO THIS SOFTWARE INCLUDING ALL IMPLIED
WARRANTIES OF MERCHANTABILITY AND FITNESS. IN NO EVENT SHALL THE
AUTHOR BE LIABLE FOR ANY SPECIAL, DIRECT, INDIRECT, OR CONSEQUENTIAL
DAMAGES OR ANY DAMAGES WHATSOEVER RESULTING FROM LOSS OF USE, DATA OR
PROFITS, WHETHER IN AN ACTION OF CONTRACT, NEGLIGENCE OR OTHER
TORTIOUS ACTION, ARISING OUT OF OR IN CONNECTION WITH THE USE OR
PERFORMANCE OF THIS SOFTWARE.
         
INSTALL ON OSX:

sudo easy_install pip
sudo pip install openpyxl==1.8.6

"""

import argparse
import csv
import sys
from openpyxl.workbook import Workbook

parser = argparse.ArgumentParser()
parser.add_argument("input_file")
parser.add_argument("-d", "--delimiter", default="\t", help="select delimiter character")
args = parser.parse_args()

if not ".csv" in args.input_file:
    sys.stderr.write("Error: File does not have the ending \".csv\".\n")
    sys.exit(2)

input_fh = open(args.input_file)
workbook = Workbook()
sheet = workbook.create_sheet(0)

for row_index, row in enumerate(
    csv.reader(open(args.input_file), delimiter=args.delimiter)):
    for col_index, col in enumerate(row):
        sheet.cell(row = row_index, column = col_index).value = col

workbook.save(open(args.input_file.replace(".csv", ".xlsx") , "w"))
